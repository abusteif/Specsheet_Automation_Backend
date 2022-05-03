from Specsheet_Automation.scripts.jira_automation import get_vendors_from_local, get_device_types_from_local, create_jira_issue
from Specsheet_Automation.classes.jira_api import JiraApi
import time
from openpyxl import load_workbook

def bulk_create_devices(excel_in, project_id=52075):
    jira = JiraApi()
    jira_token = jira.get_cookies().get_dict()["crowd.token_key"]
    results = []
    # We have to use excel reader (instead of csv) as the data contains commas and new line characters
    workbook_in = load_workbook(filename=excel_in)
    ws_in = workbook_in["devices"]

    for i in range(1, ws_in.max_row):
    # for i in range(94, 95):

        vendor = ws_in["A"+str(i)].value
        device_type = ws_in["B"+str(i)].value
        model = str(ws_in["C"+str(i)].value)
        market_name = ws_in["D"+str(i)].value
        if vendor in [
            # "Apple",
            "ACME Device Register",
            "NDA"
        ]:
            continue
        if device_type in ["LTE Modules", "Integrated Device", "5G Modules"]:
            market_name = ""
        if market_name:
            market_name = market_name.replace("\n", "")
        if model:
            model = model.replace("\n", "")
        summary = "{} {} ({})".format(vendor, model, market_name)
        request_data = {
                "vendor": vendor,
                "projectId": project_id,
                "issueType": "device",
                "summary": summary
            }
        # print(request_data)
        response = create_jira_issue("device", request_data, None, jira_token)
        # response = [True,{"text": {"key": str(i)}}]
        if response[0]:
            ws_in["E"+str(i)] = "https://jira.tools.telstra.com/browse/" + response[1]["text"]["key"]
        else:
            ws_in["E"+str(i)] = "failed to upload"
            print("failed to upload " + str(i))
        ws_in["E"+str(i)].style = "Hyperlink"
        workbook_in.save(excel_in)
        time.sleep(5)

def create_vendor_list(excel_doc, project_id=52075):
    vendors = get_vendors_from_local(project_id)
    with open(excel_doc, "w") as c:
        c.write("name,id\n")
        for v in vendors[1]:
            c.write("{},{}\n".format(v["name"], v["id"]))

def create_device_type_list(excel_doc, project_id=52075):
    device_types = get_device_types_from_local(project_id)
    with open(excel_doc, "w") as c:
        c.write("type,id\n")
        for v in device_types[1]:
            c.write("{},{}\n".format(v["name"], v["id"]))


# create_vendor_list("vendor_list.csv")
# create_device_type_list("device_type_list.csv")
bulk_create_devices("device list.xlsx")
