from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from shutil import copyfile
from copy import deepcopy, copy
from Specsheet_Automation.classes.NR_UECI_dut_info_extraction import NRDutUECIInfoExtraction
from Specsheet_Automation.classes.LTE_UECI_dut_info_extraction import LTEDutUECIInfoExtraction
from Specsheet_Automation.static_data.file_info import NR_LTE_specsheet_template
from Specsheet_Automation.classes.DUT_spec_attach_request_info_extraction import DUTSpecAttachRequestInfoExtraction
from Specsheet_Automation.static_data.LTE_specsheet_fields import MSR0835_all_UECI_fields, LTE_attach_request_fields
from Specsheet_Automation.static_data.NR_specsheet_fields import NR_all_UECI_fields, NR_attach_request_fields, \
    NR_band_fields
from Specsheet_Automation.classes.NR_data_analysis import NRDataAnalysis
from Specsheet_Automation.classes.LTE_data_analysis import LTEDataAnalysis

def populate_NR_LTE_specsheet(NR_LTE_specsheet_full_path, UECapabilityInfo_lists_file=None,
                              NR_UECapabilityInfo_lists_file=None, attach_request_lists_file=None):
    try:
        print("Populating LTE and ENDC Spec sheets")
        full_path = NR_LTE_specsheet_full_path
        copyfile(NR_LTE_specsheet_template, full_path)
        workbook = load_workbook(filename=full_path)
        lte = {
            "elements_sheet": workbook["LTE"],
            "band_combinations_sheet": workbook["LTE Band Combinations"],
            "all_fields": deepcopy(MSR0835_all_UECI_fields),
            "data": None,
            "ie_list": None,
            "ie_non_list": None,
            "data_analysis": None,
            "band_combinations": None,
            "attach_request_fields": LTE_attach_request_fields
        }

        rats = {
            "LTE": lte,
        }
        if NR_UECapabilityInfo_lists_file:
            nr = {
                "elements_sheet": workbook["ENDC"],
                "band_combinations_sheet": workbook["ENDC Band Combinations"],
                "all_fields": deepcopy(NR_all_UECI_fields),
                "data": None,
                "ie_list": None,
                "ie_non_list": None,
                "data_analysis": None,
                "band_combinations": None,
                "band_details": None,
                "ie_column": "B",
                "value_column": "C",
                "ie_ref_cell": "B35",
                "value_ref_cell": "C35",
                "attach_request_fields": NR_attach_request_fields

            }
            rats["NR"] = nr

        else:
            ENDC_sheet = workbook.get_sheet_by_name('ENDC')
            ENDC_band_combinations_sheet = workbook.get_sheet_by_name('ENDC Band Combinations')

            workbook.remove_sheet(ENDC_sheet)
            workbook.remove_sheet(ENDC_band_combinations_sheet)

        old_cell = ""
        old_result = ""
        full_result = ""

        rats["LTE"]["data"] = LTEDutUECIInfoExtraction(UECapabilityInfo_lists_file)
        rats["LTE"]["ie_list"] = rats["LTE"]["data"].ie_list
        rats["LTE"]["ie_non_list"] = rats["LTE"]["data"].ie_non_list
        rats["LTE"]["data_analysis"] = LTEDataAnalysis({**rats["LTE"]["ie_list"], **rats["LTE"]["ie_non_list"]})
        rats["LTE"]["data_analysis"].get_r10_band_combinations()
        rats["LTE"]["data_analysis"].get_r11_band_combinations()
        rats["LTE"]["band_combinations"] = rats["LTE"]["data_analysis"].band_combinations_table()

        if "NR" in rats:
            rats["NR"]["data"] = NRDutUECIInfoExtraction(NR_UECapabilityInfo_lists_file)
            rats["NR"]["ie_list"] = rats["NR"]["data"].ie_list
            rats["NR"]["ie_non_list"] = rats["NR"]["data"].ie_non_list
            rats["NR"]["data_analysis"] = NRDataAnalysis({**rats["NR"]["ie_list"], **rats["NR"]["ie_non_list"]},
                                                         {**rats["LTE"]["ie_list"], **rats["LTE"]["ie_non_list"]})
            rats["NR"]["data_analysis"].get_band_combinations()
            rats["NR"]["data_analysis"].get_supported_NR_band_details()
            rats["NR"]["band_details"] = rats["NR"]["data_analysis"].NR_bands
            rats["NR"]["band_combinations"] = rats["NR"]["data_analysis"].band_combinations_table()

        for rat in rats:
            for cell in rats[rat]["band_combinations"]:
                rats[rat]["band_combinations_sheet"][cell] = rats[rat]["band_combinations"][cell]
            for ie in rats[rat]["all_fields"]:
                if "processor" in list(ie.keys()):
                    result = rats[rat]["data_analysis"].processors[ie["processor"]]()
                    if "postProcessor" in list(ie.keys()):

                        result = ie["postProcessor"](result)
                    rats[rat]["elements_sheet"][ie["cell"]] = result
                else:
                    result = rats[rat]["data"].find_ie(ie["path"], ie["release"])
                    if not result:
                        try:
                            result = ie["default"]
                        except KeyError as e:
                            return False, "Error while populating Specsheet: {}".format(repr(e))
                    else:
                        if "values" in ie.keys():
                            result = ie["values"][int(result)]
                        elif "func" in ie.keys():
                            result = ie["func"](result)
                    if ie["cell"] == old_cell:
                        if old_result.__len__() > 0:
                            full_result = "{}\n{}\n{},".format(full_result, old_result, result)
                            old_result = ""
                        else:
                            full_result = "{}\n{},".format(full_result, result)
                    else:
                        if full_result.__len__() > 0:
                            rats[rat]["elements_sheet"][old_cell] = full_result[:-1]
                            rats[rat]["elements_sheet"][ie["cell"]] = result
                            old_cell = ie["cell"]
                            full_result = ""
                        else:
                            rats[rat]["elements_sheet"][ie["cell"]] = result
                            old_cell = ie["cell"]
                            old_result = result
            if rat == "NR":
                row_counter = rats[rat]["elements_sheet"].max_row
                ie_ref_cell = rats[rat]["elements_sheet"][rats[rat]["ie_ref_cell"]]
                value_ref_cell = rats[rat]["elements_sheet"][rats[rat]["value_ref_cell"]]
                ie_column = rats[rat]["ie_column"]
                value_column = rats[rat]["value_column"]
                for band_index, _ in enumerate(rats[rat]["band_details"]):
                    for ie in NR_band_fields:

                        row_counter += 1
                        rats[rat]["elements_sheet"]["{}{}".format(ie_column, str(row_counter))] = ie
                        apply_styling(ie_ref_cell, rats[rat]["elements_sheet"]["{}{}".format(ie_column,
                                                                                             str(row_counter))])

                        rats[rat]["elements_sheet"]["{}{}".format(value_column, str(row_counter))] = \
                            rats[rat]["band_details"][band_index][ie]
                        apply_styling(value_ref_cell, rats[rat]["elements_sheet"]["{}{}".
                                      format(value_column, str(row_counter))])

                        if ie == "NR Band":
                            thick = Side(border_style="thick")
                            thin = Side(border_style="thin")
                            rats[rat]["elements_sheet"]["{}{}".format(ie_column, str(row_counter))].\
                                border = Border(top=thick, right=thin, left=thin, bottom=thin)
                            rats[rat]["elements_sheet"]["{}{}".format(value_column, str(row_counter))].\
                                border = Border(top=thick, right=thin, left=thin, bottom=thin)
            if attach_request_lists_file:
                attach_request = DUTSpecAttachRequestInfoExtraction(attach_request_lists_file)
                for ie in rats[rat]["attach_request_fields"]:
                    result = attach_request.find_ie(ie["element"])
                    if not result:
                        try:
                            result = ie["default"]
                            rats[rat]["elements_sheet"][ie["cell"]] = result
                        except KeyError as e:
                            print("Mandatory Attach Request element not found: {}".format(ie))
                            return False, "Mandatory Attach Request element not found: {}".format(ie)
                    else:
                        rats[rat]["elements_sheet"][ie["cell"]] = ie["values"][int(result)]
        workbook.save(filename=full_path)
        assert False
    except Exception as e:
        print(repr(e))
        return False, "Error while populating Specsheet: {}".format(repr(e))
    return True, "Successfully populated Specsheet"

def apply_styling(ref_cell, current_cell):
    current_cell.font = copy(ref_cell.font)
    current_cell.border = copy(ref_cell.border)
    current_cell.fill = copy(ref_cell.fill)
    current_cell.alignment = copy(ref_cell.alignment)